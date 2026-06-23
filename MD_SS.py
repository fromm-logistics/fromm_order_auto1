# MD_SS.py
import streamlit as st
import pandas as pd
import io, re, random
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ───────────────────────────────────────────────────
# 1) 유저가 수정·확장 가능한 영역: 
#    – 기본 target_products(재고명→무게 매핑)
#    – box_limit(박스 최대 용량)
# ───────────────────────────────────────────────────
target_products={
    "[이태빈 팬미팅_PIT-A-PAT] 여권&포토카드": 200,
    "[이태빈 팬미팅_PIT-A-PAT] 마그넷": 200,
    "[이태빈 팬미팅_PIT-A-PAT] 아크릴 응원봉": 1000,
    "[이태빈 팬미팅_PIT-A-PAT] 러기지택": 200,
    "[김재중 콘서트MD_FLOWER] POSTCARD SET": 700,
    "[김재중 콘서트MD_FLOWER] T-SHIRT / WHITE": 1250,
    "[김재중 콘서트MD_FLOWER] T-SHIRT / BLACK": 1250,
    "[김재중 콘서트MD_IM TWENTY] LIGHT STICK BONNET": 200,
    "[김재중 콘서트MD_IM TWENTY] HOODIE": 1875,
    "[김재중 콘서트MD_IM TWENTY] MINI LIGHT STICK KEYRING": 500,
    "[김재중 콘서트MD_IM TWENTY] ANTI-RADIATION STICKER": 200,
    "[김재중 콘서트MD_IM TWENTY] ACRYLIC STAND A": 500,
    "[김재중 콘서트MD_IM TWENTY] ACRYLIC STAND B": 500,
    "[김재중 콘서트MD_IM TWENTY] POWER BANK": 500,
    "[김재중 콘서트MD_IM TWENTY] ECO BAG": 1000,
    "[김재중 콘서트MD_IM TWENTY] IMAGE PICKET": 1000,
    "[김재중 콘서트MD_IM TWENTY] PHOTO BLANKET": 1500,
    "[김재중 콘서트MD_IM TWENTY] RING HOLDER": 500,
    "[김재중 콘서트MD_IM TWENTY] MINI LIGHT STICK PEN": 500,
    "[김재중 콘서트MD_IM TWENTY] LIGHTSTICK CASE": 2000,
    "[김재중 콘서트MD_IM TWENTY] GRIP HOLDER": 500,
    "[김재중] OFFICIAL LIGHT STICK": 2500,
    "[김재중 콘서트MD_FLOWER] TUMBLER": 2000,
    "[김재중 콘서트MD_FLOWER] DISPOSABLE CAMERA": 1000,
    "[Shelter Of Dreams] 맨투맨": 1875,
    "[온리원오브] OFFICIAL LIGHT STICK": 1875,
    "[이승윤 콘서트MD] OFFICIAL LIGHT STICK": 1875,
    "[백현 콘서트MD_LONSDALEITE] ACRYLIC KEYRING": 500,
    "[백현 콘서트MD_LONSDALEITE] TRADING CARD": 200,
    "[백현 콘서트MD_LONSDALEITE] PHOTO JOURNAL": 500,
    "[백현 콘서트MD_LONSDALEITE] TINCASE STICKER SET": 300,
    "[백현 콘서트MD_LONSDALEITE] SHORT SLEEVED SHIRT / WHITE": 750,
    "[백현 콘서트MD_LONSDALEITE] SHORT SLEEVED SHIRT / BLACK": 750,
    "[백현 콘서트MD_LONSDALEITE] HOODED SWEATSHIRT": 1750,
    "[백현 콘서트MD_LONSDALEITE] LIGHT STICK POUCH": 1750,
    "[백현 콘서트MD_LONSDALEITE] LIGHT STICK POUCH PHOTOCARD": 1,
    "[백현 콘서트MD_LONSDALEITE] CAMP CAP": 1000,
    "[백현 콘서트MD_LONSDALEITE] CAMP CAP PHOTOCARD": 1,
    "[백현 콘서트MD_LONSDALEITE] MINI POSTER SET": 1000,
    "[백현 콘서트MD_LONSDALEITE] POSTCARD SET / LIGHT VER.": 1000,
    "[백현 콘서트MD_LONSDALEITE] POSTCARD SET / LONSDALEITE VER.": 1000,
    "[온유_FLOW] 스티커 팩": 200,
    "[ONEW FLOW] 클리어 파우치" : 500,
    "[김재중 콘서트MD_FLOWER] BANDANA / WHITE": 1000,
    "[김재중 콘서트MD_FLOWER] BANDANA / BLACK": 1000,
    "[김재중 콘서트MD_FLOWER] POUCH + KEYRING SET / GREEN": 500,
    "[김재중 콘서트MD_FLOWER] T-SHIRT / WHITE": 1000,
    "[FLOWER GARDEN] POUCH + KEYRING SET_BLACK" : 500,
    "[김재중 콘서트MD_IM TWENTY] LIGHTSTICK CASE": 2500,
    "[김재중 콘서트MD_FLOWER] SLOGAN": 1000,
    "[이승윤 콘서트MD] OFFICIAL LIGHT STICK" : 2500,
    "[휘인 앵콜콘MD_BEYOND] 볼캡 포토카드 (1종)": 1,
    "이승윤 공식응원봉 파우치": 1875,
    "[이승윤 콘서트_역성MD] 리유저블백": 2500,
    "[이승윤 콘서트_역성MD] 미니포스터세트": 1000,
    "[이승윤 콘서트_역성MD] 노트": 500,
    "[이승윤 콘서트_역성MD] 보조배터리": 1000,
    "[백현 콘서트MD_LONSDALEITE] CAMP CAP": 1000,
    "[이승윤_OFFICIAL 슬로건] OFFICAL SLOGAN": 1000,
    "[ONEW FLOW] 머그컵" : 1000,
    "[온유_FLOW] 펜 메모패드 세트": 200,
    "[온유_FLOW] 트레이딩 포토카드": 200,
    "[온유_FLOW] 미니 포스터 - BEAT VER" : 500,
    "[온유_FLOW] 미니 포스터 - DRUM VER" : 500,
    "[온유_FLOW] 러그" : 2500,
    "[온유_FLOW] 리유저블 백": 1000,
    "[이승윤 콘서트MD] OFFICIAL LIGHT STICK POUCH" : 2500,
    "[이승윤 콘서트 MD_뒤끝 Shelter] 금속 키링": 500,
    "[이승윤 Seasonal Break] 시즌그리팅": 2000,
    "[이승윤_팝업_YEOK SEONG STREET] 패딩조끼": 2000,
    "[이승윤_팝업_YEOK SEONG STREET] 러그": 2000,
    "[이승윤_팝업_YEOK SEONG STREET] 슬리퍼_M": 1000,
    "[이승윤_팝업_YEOK SEONG STREET] 슬리퍼_L": 1000,
    "[온유_생일파티 MD] [배송] 볼캡" : 1000,
    "[온유_생일파티 MD] [배송] 맨투맨" : 1875,
    "[온유_생일파티 MD] [배송] 볼캡 포토카드" : 1,
    "[온유_FLOW] 아크릴 키링 - YELLOW": 200,
    "[온유_FLOW] 아크릴 키링 - PURPLE": 200,
    "[온유_FLOW] 아크릴 키링 - PINK": 200,
    "[온유_FLOW] 담요": 2000,
    "[온유_FLOW] 스티커 세트" : 200,
    "[온유_FLOW] 스마트톡 - RABBIT VER" : 300,
    "[온유_FLOW] 스마트톡 - HEART VER" : 300,
    "[온유_생일파티 MD] [배송] 엽서 세트" : 500,
    "[온유_생일파티 MD] [배송] 맨투맨": 1750,
    "STICKER SET": 200,
    "[이승윤 콘서트_역성MD] 노트": 200,
    "[이승윤_앵콜콘_역성] 리무버블스티커" : 100,
    "[이승윤_앵콜콘_역성] 금속 뱃지" : 200,
    "[이승윤_앵콜콘_역성] 캔버스 캘린더" : 1000,
    "[다비치 콘서트MD_A Stitch] 맨투맨 - 버건디": 1750,
    "[다비치 콘서트MD_A Stitch] 맨투맨 - 그레이": 1750,
    "[다비치 콘서트MD_A Stitch] 맨투맨 - 베이지": 1750,
    "[다비치 콘서트MD_A Stitch] DAVICHI [STITCH] VINYL" : 2000,
    "[김재중_20주년_SPECIAL MD] J Ring" : 1500,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 WHITE ver_L": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 BLACK ver_L": 750,
    "[이승윤 2025 OFFICIAL MD] 민소매": 500,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 WHITE ver_M": 750,
    "[이승윤 2025 OFFICIAL MD] 캔뱃지세트 (8EA)": 100,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 GRAY ver_M": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 BLACK ver_M": 750,
    "[이승윤 2025 OFFICIAL MD] 스트링백": 1250,
    "[이승윤 2025 OFFICIAL MD] 멀티피크파우치": 300,
    "[이승윤 2025 OFFICIAL MD] 캠프캡": 500,
    "[시라카미우즈 CD] 시라카미 우즈 정규 [HAEILO] CD": 1000,
    "[이승윤 2025 OFFICIAL MD] 반다나" : 400,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 BLACK ver_2XL" : 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 BLACK ver_XL": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 GRAY ver_2XL": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 GRAY ver_L": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 GRAY ver_XL": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 WHITE ver_2XL": 750,
    "[이승윤 2025 OFFICIAL MD] 반팔티셔츠 WHITE ver_XL": 750,
    "[CAM 시라카미우즈 CD] 시라카미 우즈 정규 [HAEILO] CD" : 1000,
    "[CAM 이디오테잎] 티셔츠 / 01 SIZE" : 750,
    "[에이티즈 콘서트 MD_IN YOUR FANTASY] 미니 파우치" : 500,
    "[에이티즈 콘서트 MD_IN YOUR FANTASY] 티셔츠 링거" : 750,
    "[에이티즈 콘서트 MD_IN YOUR FANTASY] 티셔츠 블랙" : 750,
    "[에이티즈 콘서트 MD_IN YOUR FANTASY] 풀패키지 포토카드 (8종 세트)" : 200,
    "[온유_생일파티 MD] 맨투맨" : 1875,
    "[온유_생일파티 MD] 볼캡" : 1000,
    "[온유_생일파티 MD] 볼캡 포토카드" : 1,
    "[온유_생일파티 MD] 엽서 세트" : 300,
    "[이승윤 콘서트MD] PHOTO SLOGAN" : 500,
    "[이승윤 콘서트_역성MD] 스카프" : 500,
    "[이승윤_팝업_YEOK SEONG STREET] 머플러" : 500,
    "[이승윤_팝업_YEOK SEONG STREET] 여권 커버" : 500,
    "[이승윤_팝업_YEOK SEONG STREET] 은 목걸이" : 200,
    "[이승윤_팝업_YEOK SEONG STREET] 키링" : 200,
    "[CAM 10cm_<4.99999999>] 트랙 리스트 티셔츠 01size" : 750,
    "[CAM 10cm_<4.99999999>] 트랙 리스트 티셔츠 02size" : 750,
    "[CAM 다비치 콘서트MD_A Stitch] DAVICHI [STITCH] VINYL_LP" : 1000,
    "[CAM 이디오테잎] 티셔츠 / 02 SIZE" : 750,
    "[CAM 오더가든 콘서트 MD] 반팔 티셔츠 1" : 750,
    "[CAM 오더가든 콘서트 MD] 반팔 티셔츠 2" : 750,
    "[CAM 오더가든 콘서트 MD] 파우치" : 500,
    "[이승윤 생일 KIT] BRITHDAY KIT" : 1000,
    "[이승윤 생일KIT] 컵홀더 2개입 OPP포장상품" : 50,
    "[CAM 10CM 단독콘서트 5.0] Centiner Slogan Towel ver.2" : 500,
    "[CAM 10CM 단독콘서트 5.0] “5lave For You” Magnet" : 500,
    "[CAM 10CM 단독콘서트 5.0] “Follow Your Steps” Reusable Bag" : 1000,
    "[CAM 10CM 단독콘서트 5.0] “Nothing’s Going On” Ringer T-Shirts(2)" : 750,
    "[이민혁 팬미팅MD_HUTAZONE] 짐색" : 500,
    "[온유 콘서트MD 퍼센트] 리유저블백" : 1000,
    "[온유 콘서트MD 퍼센트] 메쉬파우치" : 500,
    "[온유 콘서트MD 퍼센트] 미니포스터 세트" : 500,
    "[온유 콘서트MD 퍼센트] 우양산" : 2500,
    "[온유 콘서트MD 퍼센트] 패션아트스티커" : 50,
    "[온유 콘서트MD 퍼센트] 폴딩방석" : 1500,
    "[온유 콘서트MD 퍼센트] 피크닉매트" : 5000,
    "[윤하 홀릭스 팬미팅MD] PVC 카드지갑 세트" : 500,
    "[윤하 홀릭스 팬미팅MD] 토끼 카라비너" : 500,
    "[CAM 바밍타이거 DVD] DVD" : 1000,
    "[CAM 바밍타이거 DVD] 볼캡" : 500,
    "[CAM 바밍타이거 DVD] 티셔츠 사이즈 1" : 750,
    "[CAM 바밍타이거 DVD] 티셔츠 사이즈 2" : 750,
    "[강다니엘 앵콜콘 MD] 링커티셔츠" : 750,
    "[강다니엘 앵콜콘 MD] 쉬폰 패브릭 포스터" : 750,
    "[강다니엘 앵콜콘 MD] 포토세트" : 100,
    "[이민혁 팬미팅MD_HUTAZONE] 스케줄러 세트" : 500,
    "[이민혁 팬미팅MD_HUTAZONE] 엽서 세트" : 100,
    "[이민혁 팬미팅MD_HUTAZONE] 키링" : 100,
    "[CAM 실리카겔_포토진] 포토진" : 750,
    "[CAM 다비치 콘서트MD_A Stitch] 맨투맨 - 버건디" : 750,
    "[CAM 이강승 dress EP] LP" : 1000,
    "[CAM 정세운] Official Slogan" : 750,
    "[이펙스 팬콘서트 MD_YOUTH] 미니파우치키링" : 100,
    "[이펙스 팬콘서트 MD_YOUTH] 아크릴머들러_뮤" : 100,
    "[이펙스 팬콘서트 MD_YOUTH] 아크릴머들러_에이든" : 100,
    "[이펙스 팬콘서트 MD_YOUTH] 아크릴머들러_위시" : 100,
    "[CAM 박문치 바보지퍼] 1st ALBUM" : 500,
    "[CAM 실리카겔 콘서트 MD_Syn. THE. Size X] Silica Gel CHEERING KIT v2" : 750,
    "[CAM 카더가든 콘서트 MD_BLUE HEART] 후드티_XL" : 750,
    "[온유 앵콜콘MD 퍼센트] 후드집업" : 2500,
    "[이승윤_OFFICIAL 슬로건] OFFICAL SLOGAN ver. 2" : 750,
    "[웬디 앙콘 WEALIVE MD] 마그네틱 카드지갑" : 500,
    "[웬디 앙콘 WEALIVE MD] 메탈 키링" : 300,
    "[웬디 앙콘 WEALIVE MD] 짐색" : 1000,
    "[웬디 앙콘 WEALIVE MD] 후드 집업" : 2500,
    "[유태양 팬미팅MD_THE ROOM 함께] 머그컵" : 1500,
    "[유태양 팬미팅MD_THE ROOM 함께] 머그컵 포토카드 (1종)" : 1,
    "[유태양 팬미팅MD_THE ROOM 함께] 위드 유 세트" : 40,
    "[유태양 팬미팅MD_THE ROOM 함께] 카라비너 키링" : 100,
    "[유태양 팬미팅MD_THE ROOM 함께] 트레이딩 포토카드" : 30,
    "[온유 팝업_TOUGH LOVE] 다이어리 내지_Daily ver." : 250,
    "[온유 팝업_TOUGH LOVE] 다이어리 내지_Mood ver." : 250,
    "[온유 팝업_TOUGH LOVE] 롤 테이프_Doodle ver." : 250,
    "[온유 팝업_TOUGH LOVE] 롤 테이프_Emotion ver." : 250,
    "[온유 팝업_TOUGH LOVE] 미니 포스터 세트" : 190,
    "[온유 팝업_TOUGH LOVE] 북 스티커" : 75,
    "[온유 팝업_TOUGH LOVE] 북스토퍼 세트" : 270,
    "[온유 팝업_TOUGH LOVE] 엽서 포토카드 세트" : 250,
    "[온유 팝업_TOUGH LOVE] 틴 마그넷_Love ver." : 170,
    "[온유 팝업_TOUGH LOVE] 틴 마그넷_Tough ver." : 170,
    "[온유 팝업_TOUGH LOVE] 페이퍼 인센스" : 180,
    "[온유 팝업_TOUGH LOVE] 페이퍼 인센스 특전 포토카드" : 50,
    "[온유 팝업_TOUGH LOVE] 프레임 홀더 세트" : 190,
    "[정세운 2026 시즌그리팅] 시즌그리팅" : 1363.6,
    "[CAM 10CM_Chapter 1] 날진 물통" : 5000,
    "[CAM 10CM_Chapter 1] 반팔 티셔츠" : 750,
    "[CAM 10CM_Chapter 1] 아크릴 마그넷 (도쿄)_대행" : 300,
    "[CAM 10CM_Chapter 1] 아크릴 마그넷 (방콕)_대행" : 300,
    "[CAM 10CM_Chapter 1] 아크릴 마그넷 (부산)" : 300,
    "[CAM 10CM_Chapter 1] 아크릴 마그넷 (싱가포르)_대행" : 300,
    "[CAM 10CM_Chapter 1] 아크릴 마그넷 (타이베이)_대행" : 300,
    "[CAM 10CM_Chapter 1] 아크릴 마그넷 (홍콩)_대행" : 300,
    "[CAM 10CM_Chapter 1] 파우치 대행" : 500,
    "[CAM 10CM_Chapter 1] 후드티셔츠 레드" : 1875,
    "[CAM 10CM_Chapter 1] 후드티셔츠 블랙" : 1875,
    "[CAM 머드 더 스튜던트 LAGEON] 볼캡" : 1000,
    "[CAM 머드 더 스튜던트 LAGEON] 티셔츠 L" : 750,
    "[CAM 머드 더 스튜던트 LAGEON] 티셔츠 M" : 750,
    "[CAM 머드 더 스튜던트 LAGEON] 포토진" : 1500,
    "[CAM 머드 더 스튜던트 LAGEON] 후드집업 L" : 1875,
    "[CAM 정세운 콘서트MD_Margins] 기타 피크 & 핀 배지 세트" : 300,
    "[CAM 정세운 콘서트MD_Margins] 슬리퍼" : 1000,
    "[CAM 정세운 콘서트MD_Margins] 엽서" : 100,
    "[CAM 정세운 콘서트MD_Margins] 타포린백" : 1500,
    "[CAM 정세운 콘서트MD_Margins] 티셔츠" : 750,
    "[CAM 정세운] Keyring (Love in the Margins Edition)" : 300,
    "[CAM 카더가든 콘서트 MD_BLUE HEART] 후드티_M" : 1875,
    "[쇼미더머니_2025 콘서트] 리무버블 스티커" : 150
}

box_limit = 15000  # 기본 박스 최대 용량

def run_md_ss():
    st.button("◀ MD 창으로 돌아가기",
              on_click=lambda: st.session_state.update(page="md_main"),
              key="back_to_md_main_from_FS")
    st.title("📋 SS 나누기")

    uploaded = st.file_uploader("▶ SS 전용 CSV 업로드", type="csv", key="SS_csv")
    if uploaded:
        df = pd.read_csv(uploaded, dtype={'우편번호': str, '전화번호': str})
        st.session_state['SS_df'] = df
        st.success(f"CSV 업로드 완료: {df.shape[0]}행")
    else:
        df = st.session_state.get('SS_df', None)

    missing = []
    if df is not None:
        missing = sorted(set(df['재고명'].dropna()) - set(target_products))
        if missing:
            st.warning("타겟에 정의되지 않은 재고명 발견:")
            for name in missing:
                st.code(f'"{name}" : ', language="python")
            st.info("위 항목을 target_products에 추가하거나 아래에서 무게를 입력하세요.")
            if st.button("검증"):
                st.session_state['SS_verified'] = True
        else:
            st.success("모든 재고명이 target_products에 포함됩니다.")
            st.session_state['SS_verified'] = True

    if st.session_state.get('SS_verified') and missing:
        st.markdown("### 누락된 재고명의 무게를 입력해주세요")
        custom = st.session_state.get('SS_custom_weights', {})
        for prod in missing:
            w = st.number_input(f"{prod} ▶ 무게입력", min_value=1, key=f"w_{prod}")
            if w:
                custom[prod] = w
        st.session_state['SS_custom_weights'] = custom

    if st.session_state.get('SS_verified') and df is not None and st.button("✅ 실행"):
        merged_tp = {**target_products, **st.session_state.get('SS_custom_weights', {})}
        buf_all, buf_dom, buf_int = _process_ss(df, merged_tp, box_limit)
        st.session_state['SS_buf_all'] = buf_all.getvalue()
        st.session_state['SS_buf_dom'] = buf_dom.getvalue()
        st.session_state['SS_buf_int'] = buf_int.getvalue()
        st.success("SS 처리 완료!")

    if 'SS_buf_all' in st.session_state:
        today = datetime.datetime.now().strftime("%y%m%d")
        st.download_button("▶ 원본 시트 다운로드",
            st.session_state['SS_buf_all'],
            file_name=f"SS_{today}_원본.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("▶ 국내 시트만 다운로드",
            st.session_state['SS_buf_dom'],
            file_name=f"SS_{today}_국내.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("▶ 해외 시트만 다운로드",
            st.session_state['SS_buf_int'],
            file_name=f"SS_{today}_해외.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _process_ss(df: pd.DataFrame, tp: dict, limit: int):
    mask_301420 = df['재고코드'] == int(301420)
    if mask_301420.any():
        for _, row in df[mask_301420].iterrows():
            new_row = row.copy()
            new_row['id'] = ""
            new_row['재고코드'] = "301484"
            new_row['상품명(내부용)'] = "[이승윤 생일KIT] 컵홀더 2개입 OPP포장상품"
            new_row['상품명'] = "[이승윤 생일KIT] 컵홀더 2개입 OPP포장상품"
            new_row['옵션명'] = "[이승윤 생일KIT] 컵홀더 2개입 OPP포장상품"
            new_row['재고명'] = "[이승윤 생일KIT] 컵홀더 2개입 OPP포장상품"
            new_row['상품금액'] = 1
            new_row['결제통화'] = "USD"
            new_row['상품무게'] = 1
            df.loc[len(df)] = new_row
        tp["[이승윤 생일KIT] 컵홀더 2개입 OPP포장상품"] = 1

    mask_301392 = df['재고코드'] == int(301392)
    if mask_301392.any():
        for _, row in df[mask_301392].iterrows():
            new_row = row.copy()
            new_row['id'] = ""
            new_row['재고코드'] = "301400"
            new_row['상품명(내부용)'] = "[온유 콘서트MD 퍼센트] 우양산 포토카드 (1종)"
            new_row['상품명'] = "[온유 콘서트MD 퍼센트] 우양산 포토카드 (1종)"
            new_row['옵션명'] = "[온유 콘서트MD 퍼센트] 우양산 포토카드 (1종)"
            new_row['재고명'] = "[온유 콘서트MD 퍼센트] 우양산 포토카드 (1종)"
            new_row['상품금액'] = 1
            new_row['결제통화'] = "USD"
            new_row['상품무게'] = 1
            df.loc[len(df)] = new_row
        tp["[온유 콘서트MD 퍼센트] 우양산 포토카드 (1종)"] = 1

    def assign_order_numbers(group):
        total_w, suffix, out = 0, 1, []
        for _, row in group.iterrows():
            weight = tp.get(row['재고명'], 0)
            qty = int(row['수량'])
            while qty > 0:
                if total_w + weight > limit:
                    suffix, total_w = suffix + 1, 0
                can = min(qty, (limit - total_w) // weight if weight > 0 else qty)
                nr = row.copy()
                nr['수량'] = can
                nr['주문번호'] = f"{row['주문번호']}-{suffix}"
                out.append(nr)
                total_w += can * weight
                qty -= can
        return out

    rows = []
    for _, grp in df.groupby('주문번호'):
        rows += assign_order_numbers(grp)
    res = pd.DataFrame(rows)

    # 2) 금액·통화·실결제금액
    res['상품금액'] = res['상품금액'].replace(0, 1)
    m = (res['국가코드']!='KR') & (res['결제통화']=='KRW')
    res.loc[m, '결제통화'] = 'USD'
    res.loc[m, '상품금액'] = res.loc[m,'상품금액']/1000
    res['실결제금액'] = res['수량']*res['상품금액']
    cols = list(res.columns)
    cols.insert(cols.index('상품금액')+1, '실결제금액')
    res = res[cols]

    # 3) 필터링 & 컬럼 추가
    res = res[res['수량']>0].copy()
    res['상품명'] = res['재고명']
    res.loc[res['국가명'] == "Japan", '국가명'] = "."

    dom  = res[res['국가코드']=='KR'].copy()
    intl = res[res['국가코드']!='KR'].copy()

    # 4) 특정 상품 복제
    def dup_special(ddf):
        mask = ddf['상품명'] == "[김재중 아시아 투어MD_J-PARTY] 핀브로치"
        dup = ddf[mask].copy()
        if not dup.empty:
            dup['id'] = ""
            dup['재고코드'] = "500194"
            dup['상품명'] = "[김재중 아시아 투어MD_J-PARTY] 핀브로치 포토카드"
            dup['옵션명'] = "PIN BROOCH PHOTOCARD"
            ddf = pd.concat([ddf, dup], ignore_index=True)
        return ddf

    dom  = dup_special(dom)
    intl = dup_special(intl)

    # 5) 전화번호 포맷
    def fmt(p):
        s=re.sub(r"\D","",str(p))
        if len(s)==10 and s[:2] in ("10","70"): s="0"+s
        if len(s)==11: return f"{s[:3]}-{s[3:7]}-{s[7:]}"
        if len(s)==10: return f"{s[:3]}-{s[3:6]}-{s[6:]}"
        return s
    dom['전화번호'] = dom['전화번호'].apply(fmt)

    # 6) 국제 컬럼 보정
    # ⭐ [조건 반영] 주석 해제 후 국가코드별 배송사 자동 매핑 규칙 적용
    intl['희망배송사'] = intl['국가코드'].map(
        lambda c: 'sagawa' if c == 'JP' else ('emspremium' if c in ['US', 'IT', 'CO', 'DE', 'SE', 'NL', 'NO', 'DK'] else 'ems')
    )

    intl['주'] = intl.apply(
        lambda r: r['국가명'] if pd.isna(r['주']) or str(r['주']).strip() == '' else r['주'], axis=1)

    intl['도시'] = intl.apply(
        lambda r: r['주'] if pd.isna(r['도시']) or str(r['도시']).strip() == '' else r['도시'], axis=1)

    mask_jp = intl['국가코드'] == 'JP'
    intl.loc[mask_jp, '도로명주소'] = (
        intl.loc[mask_jp, '도로명주소'].fillna('') + ' ' +
        intl.loc[mask_jp, '상세주소'].fillna('') + ' ' +
        intl.loc[mask_jp, '도시'].fillna('') + ' ' +
        intl.loc[mask_jp, '주'].fillna('')
    ).str.strip()

    intl.loc[mask_jp, ['상세주소', '주']] = ''
    intl.loc[mask_jp, ['도시']] = '.'

    # 7) Excel 쓰기 & 하이라이트
    buf_all = io.BytesIO()
    with pd.ExcelWriter(buf_all, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="원본", index=False)
        dom.to_excel(w, sheet_name="국내", index=False)
        intl.to_excel(w, sheet_name="해외", index=False)
        inv = (
            pd.concat([dom[['상품명','수량']], intl[['상품명','수량']]])
              .groupby("상품명", as_index=False)["수량"]
              .sum().rename(columns={"상품명":"이름","수량":"갯수"})
        )
        inv.to_excel(w, sheet_name="재고_값", index=False)
    buf_all.seek(0)

    wb     = load_workbook(buf_all)
    yellow = PatternFill("solid", fgColor="FFFF00")

    ws_d = wb["국내"]
    idx_z = dom.columns.get_loc("우편번호")+1
    idx_p = dom.columns.get_loc("전화번호")+1
    for i,row in enumerate(dom.itertuples(), start=2):
        if len(str(row.우편번호))!=5:      ws_d.cell(i,idx_z).fill=yellow
        if not str(row.전화번호).startswith("010"): ws_d.cell(i,idx_p).fill=yellow

    ws_i = wb["해외"]
    c_name = intl.columns.get_loc("받는사람")+1
    c_hk   = intl.columns.get_loc("국가코드")+1
    c_zip  = intl.columns.get_loc("우편번호")+1
    for i,row in enumerate(intl.itertuples(), start=2):
        if str(row.받는사람).isdigit(): ws_i.cell(i,c_name).fill=yellow
        if row.국가코드=="HK":
            ws_i.cell(i,c_hk).fill=yellow
            if str(row.우편번호) in {"-","0","00","000","0000","00000"}:
                ws_i.cell(i,c_zip).value="999077"

    buf_all = io.BytesIO()
    wb.save(buf_all)
    buf_all.seek(0)

    buf_dom = io.BytesIO(); buf_int = io.BytesIO()
    with pd.ExcelWriter(buf_dom, engine="openpyxl") as w2:
        dom.to_excel(w2, sheet_name="국내", index=False)
    with pd.ExcelWriter(buf_int, engine="openpyxl") as w3:
        intl.to_excel(w3, sheet_name="해외", index=False)
    buf_dom.seek(0); buf_int.seek(0)

    return buf_all, buf_dom, buf_int
