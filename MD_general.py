import streamlit as st
import pandas as pd
import io, re, random
import datetime
import math
import numpy as np  
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 가격 예측 및 카테고리 설정   
def _process_general(df: pd.DataFrame, weight_map: dict, box_limit: int):
    df = df.copy()
    df['수량'] = pd.to_numeric(df['수량'], errors='coerce').fillna(0).astype(int)  # ← 추가

    def assign_boxes(group):
        total_weight = 0
        order_suffix = 1
        result_rows = []

        for index, row in group.iterrows():
            product_name = row['상품명']
            weight = weight_map.get(product_name, 0)
            qty = int(row['수량'])

            while qty > 0:
                if total_weight + weight > box_limit:
                    order_suffix += 1
                    total_weight = 0

                max_packable = (box_limit - total_weight) / weight if weight > 0 else qty
                can_pack = min(qty, math.floor(max_packable))
                if can_pack <= 0:  
                    can_pack = 1

                new_row = row.copy()
                new_row['수량'] = can_pack
                new_row['주문번호'] = f"{row['주문번호']}-{order_suffix}"
                result_rows.append(new_row)

                total_weight += can_pack * weight
                qty -= can_pack

        return result_rows

    rows = []
    for _, grp in df.groupby('주문번호'):
        rows += assign_boxes(grp)
    res = pd.DataFrame(rows)

    # 2) 금액·통화·실결제금액
    res['상품금액'] = pd.to_numeric(res['상품금액'], errors='coerce').fillna(1)
    res['상품금액'] = res['상품금액'].replace(0, 1)
    m = (res['국가코드']!='KR') & (res['결제통화']=='KRW')
    res.loc[m, '결제통화'] = 'USD'
    res.loc[m, '상품금액'] = res.loc[m,'상품금액']/1000
    res['실결제금액'] = res['수량']*res['상품금액']
    cols = list(res.columns)
    cols.insert(cols.index('상품금액')+1, '실결제금액')
    res = res[cols]

    # 3) 필터링 & 컬럼 추가
    res = res[res['수량']>0].copy()
    res['상품명'] = res['상품명']
    res.loc[res['국가명'] == "Japan", '국가명'] = "."

    dom  = res[res['국가코드']=='KR'].copy()
    intl = res[res['국가코드']!='KR'].copy()

    # 4) 특정 상품 복제
    def dup_special(ddf):
        mask = ddf['상품명'] == "[김재중 아시아 투어MD_J-PARTY] 핀브로치"
        dup = ddf[mask].copy()
        if not dup.empty:
            dup['id'] = ""
            dup['재고코드'] = "500194"
            dup['상품명'] = "[김재중 아시아 투어MD_J-PARTY] 핀브로치 포토카드"
            dup['옵션명'] = "PIN BROOCH PHOTOCARD"
            ddf = pd.concat([ddf, dup], ignore_index=True)
        return ddf

    dom  = dup_special(dom)
    intl = dup_special(intl)

    # 5) 전화번호 포맷
    def fmt(p):
        s=re.sub(r"\D","",str(p))
        if len(s)==10 and s[:2] in ("10","70"): s="0"+s
        if len(s)==11: return f"{s[:3]}-{s[3:7]}-{s[7:]}"
        if len(s)==10: return f"{s[:3]}-{s[3:6]}-{s[6:]}"
        return s
    dom['전화번호'] = dom['전화번호'].apply(fmt)

    # 6) 국제 컬럼 보정
    # ⭐ [조건 반영] 국가코드별 희망배송사 매핑 조건 확장 규칙 적용
    intl['희망배송사'] = intl['국가코드'].map(
        lambda c: 'sagawa' if c == 'JP' else ('emspremium' if c in ['US', 'IT', 'CO', 'RO'] else 'ems')
    )

    intl['주'] = intl.apply(
        lambda r: r['국가명'] if pd.isna(r['주']) or str(r['주']).strip() == '' else r['주'], axis=1)

    intl['도시'] = intl.apply(
        lambda r: r['주'] if pd.isna(r['도시']) or str(r['도시']).strip() == '' else r['도시'], axis=1)
    
    mask_jp = intl['국가코드'] == 'JP'
    intl.loc[mask_jp, '도로명주소'] = (
        intl.loc[mask_jp, '도로명주소'].fillna('') + ' ' +
        intl.loc[mask_jp, '상세주소'].fillna('') + ' ' +
        intl.loc[mask_jp, '도시'].fillna('') + ' ' +
        intl.loc[mask_jp, '주'].fillna('')
    ).str.strip()

    intl.loc[mask_jp, ['상세주소', '주']] = ''
    intl.loc[mask_jp, ['도시']] = '.'

    # 7) 판매사 품주번호
    def gen(r,i):
        if r['상품명'].endswith("핀브로치 포토카드"):
            return str(random.randint(10**14,10**15-1))
        if not r['id']:
            rnd = f"{random.randint(0,9999999999):010d}"
            return f"NAN_{rnd[:3]}_{rnd[3:6]}_{rnd[6:]}"
        return f"{r['id']}-{str(r.get('배송비묶음',''))[:3]}-{i+1}"

    intl['판매사 품주번호'] = [gen(row,idx) for idx,row in intl.iterrows()]
    seen={}
    def uniquify(vals):
        out=[]
        for v in vals:
            if v in seen: seen[v]+=1; out.append(f"{v}-{seen[v]}")
            else: seen[v]=0; out.append(v)
        return out

    intl['판매사 품주번호'] = uniquify(intl['판매사 품주번호'])

    # 8) Excel 쓰기 & 하이라이트
    buf_all = io.BytesIO()
    with pd.ExcelWriter(buf_all, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="원본", index=False)
        dom.to_excel(w, sheet_name="국내", index=False)
        intl.to_excel(w, sheet_name="해외", index=False)
        inv = (
            pd.concat([dom[['상품명','수량']], intl[['상품명','수량']]])
            .groupby("상품명", as_index=False)["수량"]
            .sum().rename(columns={"상품명":"이름","수량":"갯수"})
        )
        inv.to_excel(w, sheet_name="재고_값", index=False)
    buf_all.seek(0)

    wb     = load_workbook(buf_all)
    yellow = PatternFill("solid", fgColor="FFFF00")

    ws_d = wb["국내"]
    idx_z = dom.columns.get_loc("우편번호")+1
    idx_p = dom.columns.get_loc("전화번호")+1
    for i,row in enumerate(dom.itertuples(), start=2):
        if len(str(row.우편번호))!=5:      ws_d.cell(i,idx_z).fill=yellow
        if not str(row.전화번호).startswith("010"): ws_d.cell(i,idx_p).fill=yellow

    ws_i = wb["해외"]
    c_name = intl.columns.get_loc("받는사람")+1
    c_hk   = intl.columns.get_loc("국가코드")+1
    c_zip  = intl.columns.get_loc("우편번호")+1
    for i,row in enumerate(intl.itertuples(), start=2):
        if str(row.받는사람).isdigit(): ws_i.cell(i,c_name).fill=yellow
        if row.국가코드=="HK":
            ws_i.cell(i,c_hk).fill=yellow
            if str(row.우편번호) in {"-","0","00","000","0000","00000"}:
                ws_i.cell(i,c_zip).value="999077"

    buf_all = io.BytesIO()
    wb.save(buf_all)
    buf_all.seek(0)

    buf_dom = io.BytesIO(); buf_int = io.BytesIO()
    with pd.ExcelWriter(buf_dom, engine="openpyxl") as w2:
        dom.to_excel(w2, sheet_name="국내", index=False)
    with pd.ExcelWriter(buf_int, engine="openpyxl") as w3:
        intl.to_excel(w3, sheet_name="해외", index=False)
    buf_dom.seek(0); buf_int.seek(0)

    return buf_all, buf_dom, buf_int

   
def run_md_general():
    st.button(
        "◀ MD 창으로 돌아가기",
        on_click=lambda: st.session_state.update(page="md_main"),
        key="back_to_md_main_from_general"
    )

    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'df_step1' not in st.session_state:
        st.session_state.df_step1 = None
    if 'df_step2' not in st.session_state:
        st.session_state.df_step2 = None

    def go_next():
            st.session_state.step += 1
    def go_prev():
            st.session_state.step = max(1, st.session_state.step - 1)

    st.title("📦 MD 자동화")

    col1, col2, col3 = st.columns([1,1,1])
    with col1:
        st.button("◀ 이전 단계", on_click=go_prev)
    with col2:
        st.markdown(f"### Step {st.session_state.step} / 3")
    with col3:
        st.button("다음 단계 ▶", on_click=go_next)

    st.write("---")

    if st.session_state.step == 1:
        st.subheader("1️⃣ 상품 나누기 진행")

        uploaded_file = st.file_uploader("▶ CSV 파일 업로드", type=["csv"], key="fs_csv_upload")
        if uploaded_file:
            df = pd.read_csv(uploaded_file, dtype=str)
            st.session_state["fs_df"] = df
            st.success(f"CSV 업로드 완료: {df.shape[0]}행")
        else:
            df = st.session_state.get("fs_df", None)

        if df is not None:
            product_names = sorted(df["상품명"].dropna().unique())
            saved_weights = st.session_state.get("fs_custom_weights", {})
            new_weights = {}

            st.markdown("### 🧾 상품명과 무게 입력")
            st.markdown("""<style>
            .product-box {
                display: inline-block;
                width: 260px;
                margin: 10px 12px 10px 0;
                padding: 10px 10px 8px 10px;
                border: 1px solid #ddd;
                border-radius: 6px;
                vertical-align: top;
            }
            .product-name {
                font-size: 14px;
                font-weight: 500;
                margin-bottom: 4px;
                word-break: break-word;
            }
            .stNumberInput {
                margin-top: -4px;
            }
            </style>""", unsafe_allow_html=True)

            for i, prod in enumerate(product_names):
                try:
                    default_weight = int(float(df[df["상품명"] == prod]["상품무게"].dropna().iloc[0]))
                except Exception:
                    default_weight = 100

                with st.container():
                    cols = st.columns([1])  
                    with cols[0]:
                        st.markdown(f"<div class='product-box'><div class='product-name'>{prod}</div>", unsafe_allow_html=True)
                        weight = st.number_input(
                            label="",
                            min_value=1,
                            value=saved_weights.get(prod, default_weight),
                            key=f"weight_{prod}",
                            label_visibility="collapsed"
                        )
                        st.markdown("</div>", unsafe_allow_html=True)
                        new_weights[prod] = weight

            st.session_state["fs_custom_weights"] = new_weights
            st.success("✅ 모든 무게 정보가 저장되었습니다.")

        if df is not None and st.button("✅ 실행"):
            weights = st.session_state.get('fs_custom_weights', {})
            buf_all, buf_dom, buf_int = _process_general(df, weights, box_limit=15000)
            st.session_state['fs_buf_all'] = buf_all.getvalue()
            st.session_state['fs_buf_dom'] = buf_dom.getvalue()
            st.session_state['fs_buf_int'] = buf_int.getvalue()
            st.success("GENERAL 변환 처리 완료!")

        if 'fs_buf_all' in st.session_state:
            today = datetime.datetime.now().strftime("%y%m%d")
            st.download_button("▶ 원본 시트 다운로드",
                st.session_state['fs_buf_all'],
                file_name=f"일반상품_{today}_원본.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.download_button("▶ 국내 시트만 다운로드",
                st.session_state['fs_buf_dom'],
                file_name=f"일반상품_{today}_국내.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.download_button("▶ 해외 시트만 다운로드",
                st.session_state['fs_buf_int'],
                file_name=f"일반상품_{today}_해외.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            

    elif st.session_state.step == 2:
        st.subheader("2️⃣ 가격 예측 및 카테고리 설정")
        def run_md_prediction():
            PRICE_TABLE_PATH = "가격테이블.xlsx"

            xlsx_file = st.file_uploader("▶ Step1 결과물 업로드 (XLSX)", type=["xlsx"], key="xlsx_upload")
            if xlsx_file:
                wb = pd.read_excel(xlsx_file, sheet_name=None)
                df_origin = wb['원본']
                df_domestic = wb['국내']
                df_stock = wb['재고_값']
                df_domestic.columns = df_domestic.columns.str.strip()
                if '상품명' not in df_domestic.columns:
                    raise KeyError("❗ 국내 시트에 '상품명' 컬럼이 없습니다. 컬럼명을 확인하세요.")
                df_international = wb['해외']
                df_international.columns = df_international.columns.str.strip()
                if '상품명' not in df_international.columns:
                    raise KeyError("❗ 해외 시트에 '상품명' 컬럼이 없습니다. 컬럼명을 확인하세요.")

                category_options = [
                    "멤버십키트","키링", "로브", "가방", "담요", "응원봉(파우치)", "맨투맨", "숄더백", "에코백", "트랙수트", "후드",
                    "머플러", "티셔츠", "쉬폰포스터", "슬로건", "작은인형", "헤드폰", "슬리퍼", "모자류",
                    "포스터", "지류", "퍼즐", "티켓홀더백", "노트", "목걸이등악세서리", "학용품",
                    "엽서", "아크릴스탠드", "컵", "파우치", "트레이딩포토카드", "양말류", "코스터",
                    "포토카드(증정)", "포토키트", "향수","CD","앨범(키위)","앨범(일반)"
                ]

                st.subheader("1️⃣ 재고명별 카테고리 선택")
                product_names = sorted(df_stock['이름'].dropna().unique())
                category_map = {}
                for prod in product_names:
                    category_map[prod] = st.selectbox(f"{prod}", category_options, key=f"cat_{prod}")

                if st.button("✅ 예측 실행"):
                    st.success("예측 처리 중...")

                    df_domestic['카테고리'] = df_domestic['재고명'].map(category_map)
                    df_international['카테고리'] = df_international['재고명'].map(category_map)

                    weight_table = pd.read_excel(PRICE_TABLE_PATH, sheet_name='MD청구무게')
                    weight_dict = dict(zip(weight_table['카테고리'], weight_table['MD청구무게']))

                    df_domestic['실청구무게'] = df_domestic['카테고리'].map(weight_dict).fillna(0).astype(int) * df_domestic['수량']
                    df_international['실청구무게'] = df_international['카테고리'].map(weight_dict).fillna(0).astype(int) * df_international['수량']

                    def summarize(df, label):
                        df['상품무게'] = pd.to_numeric(df['상품무게'], errors='coerce').fillna(0) * df['수량']
                        summary = df.groupby('주문번호').agg({
                            '상품무게': 'sum',
                            '실청구무게': 'sum',
                            '수량': 'sum',
                            '국가코드': 'first'
                        }).reset_index()
                        summary['보정청구무게'] = summary['실청구무게'].apply(lambda x: int(math.ceil(x / 500.0)) * 500)
                        return summary
                    
                    dom_pivot = summarize(df_domestic.copy(), '국내')
                    int_pivot = summarize(df_international.copy(), '해외')

                    price_table = pd.read_excel(PRICE_TABLE_PATH, sheet_name='해외배송백데이터')
                    price_table['무게'] = price_table['무게'].astype(int)

                    def find_prices(row):
                        matched = price_table[(price_table['국가코드'] == row['국가코드']) & (price_table['무게'] == row['보정청구무게'])]
                        if matched.empty:
                            return pd.Series([np.nan, np.nan])
                        return pd.Series([
                            matched.iloc[0]['다보내배송비'],
                            matched.iloc[0]['패박배송비']
                        ])

                    int_pivot[['다보내 예상 금액', '패박 예상 금액']] = int_pivot.apply(find_prices, axis=1)

                    def calc_labors(qty):
                        dabo = 1000 + min(1000, max(0, (qty - 4) * 50)) + 100
                        if qty >= 5:
                            fast = 1200 + (qty - 4) * 70 + 150
                        else:
                            fast = 1200 + 150
                        return pd.Series([dabo, fast])

                    int_pivot[['다보내 작업비', '패스트박스 작업비']] = int_pivot['수량'].apply(calc_labors)
                    int_pivot['다보내 총 배송비'] = int_pivot['다보내 예상 금액'] + int_pivot['다보내 작업비']
                    int_pivot['패스트박스 총 배송비'] = int_pivot['패박 예상 금액'] + int_pivot['패스트박스 작업비']

                    def add_total(df):
                        numeric_cols = df.select_dtypes(include=[np.number]).columns
                        total = df.loc[~df['주문번호'].eq('총 합계'), numeric_cols].sum()
                        total['주문번호'] = '총 합계'
                        return pd.concat([df, pd.DataFrame([total])], ignore_index=True)
                    
                    dom_pivot[['다보내 예상 금액', '패박 예상 금액']] = dom_pivot.apply(lambda row: pd.Series([
                        2300 + 1000 + min(1000, max(0, (row['수량'] - 4) * 50)) + 100,
                        2100 + 1100 + ((row['수량'] - 4) * 70 + 150 if row['수량'] >= 5 else 150)
                    ]), axis=1)

                    dom_pivot = add_total(dom_pivot)
                    int_pivot = add_total(int_pivot)

                    dom_pivot['고객수령금액'] = dom_pivot['상품무게'].apply(lambda w: math.ceil(w / 15000) * 3000)

                    admin_price_table = pd.read_excel(PRICE_TABLE_PATH, sheet_name='프롬어드민가격테이블')
                    admin_price_table['무게'] = admin_price_table['무게'].astype(int)

                    def get_admin_price(row):
                        weight = row['상품무게']
                        country = row['국가코드']

                        if weight <= 15000:
                            rounded = math.ceil(weight / 1000) * 1000
                            match = admin_price_table[
                                (admin_price_table['국가코드'] == country) &
                                (admin_price_table['무게'] == rounded)
                            ]
                            if not match.empty:
                                return match.iloc[0]['수령금액']
                            return 0
                        else:
                            base_match = admin_price_table[
                                (admin_price_table['국가코드'] == country) &
                                (admin_price_table['무게'] == 15000)
                            ]
                            base_price = base_match.iloc[0]['수령금액'] if not base_match.empty else 0
                            excess_weight = weight - 1000
                            extra_unit_count = math.ceil(excess_weight / 1000)
                            unit_match = admin_price_table[
                                (admin_price_table['국가코드'] == country) &
                                (admin_price_table['무게'] == 1000)
                            ]
                            unit_price = unit_match.iloc[0]['수령금액'] if not unit_match.empty else 0
                            return base_price + extra_unit_count * unit_price

                    int_pivot['고객수령금액'] = int_pivot.apply(get_admin_price, axis=1)
                    int_pivot['고객수령금액'] *= 1.3 
                    
                    total_dom_rows = len(dom_pivot[dom_pivot['주문번호'] != '총 합계'])
                    total_int_rows = len(int_pivot[int_pivot['주문번호'] != '총 합계'])
                    total_rows = total_dom_rows + total_int_rows

                    ratio_international = total_int_rows / total_rows if total_rows else 0
                    ratio_japan = len(int_pivot[(int_pivot['국가코드'] == 'JP') & (int_pivot['주문번호'] != '총 합계')]) / total_rows if total_rows else 0

                    dabo_domestic_sum = dom_pivot[dom_pivot['주문번호'] != '총 합계']['다보내 예상 금액'].sum()
                    dabo_international_sum = int_pivot[int_pivot['주문번호'] != '총 합계']['다보내 총 배송비'].sum()

                    fast_domestic_sum = dom_pivot[dom_pivot['주문번호'] != '총 합계']['패박 예상 금액'].sum()
                    fast_international_sum = int_pivot[int_pivot['주문번호'] != '총 합계']['패스트박스 총 배송비'].sum()

                    diff = abs(dabo_domestic_sum + dabo_international_sum - fast_domestic_sum - fast_international_sum)
                    recommendation = "다보내" if (dabo_domestic_sum + dabo_international_sum) < (fast_domestic_sum + fast_international_sum) else "패스트박스"

                    total_customer_amount = dom_pivot[dom_pivot['주문번호'] != '총 합계']['고객수령금액'].sum() + \
                        int_pivot[int_pivot['주문번호'] != '총 합계']['고객수령금액'].sum()

                    domestic_international_ratio = f"{ratio_international * 100:.1f}%"
                    japan_ratio = f"{ratio_japan * 100:.1f}%"
                    prediction_summary = pd.DataFrame([{
                        "국내/해외비율": domestic_international_ratio,
                        "일본 비율": japan_ratio,
                        "다보내_국내예측": int(dabo_domestic_sum),
                        "다보내_해외 예측": int(dabo_international_sum),
                        "패박_국내예측": int(fast_domestic_sum),
                        "패박_해외예측": int(fast_international_sum),
                        "예측차이가격": int(diff),
                        "추천물류사": recommendation,
                        "고객수령금액": int(total_customer_amount)
                    }])
                    prediction_summary["다보내 총합"] = int(dabo_domestic_sum + dabo_international_sum)
                    prediction_summary["패박 총합"] = int(fast_domestic_sum + fast_international_sum)

                    lowest_cost = min(int(dabo_domestic_sum + dabo_international_sum), int(fast_domestic_sum + fast_international_sum))
                    recommendation_pl = int(total_customer_amount) - lowest_cost
                    prediction_summary["추천물류사P/L"] = recommendation_pl
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        prediction_summary.to_excel(writer, sheet_name='예측금액', index=False, startcol=1, startrow=1)
                        ws = writer.book["예측금액"]
                        ws['B5'] = "다보내 총합"
                        ws['C5'] = "패박 총합"
                        ws['B6'] = int(dabo_domestic_sum + dabo_international_sum)
                        ws['C6'] = int(fast_domestic_sum + fast_international_sum)

                        wb['예측'] = None  
                        df_origin.to_excel(writer, sheet_name='원본', index=False)
                        df_domestic.to_excel(writer, sheet_name='국내_주문서', index=False)
                        df_international.to_excel(writer, sheet_name='해외_주문서', index=False)
                        dom_pivot.to_excel(writer, sheet_name='국내_가격', index=False)
                        int_pivot.to_excel(writer, sheet_name='해외_가격', index=False)
                        inventory = df_origin.groupby('상품명')['수량'].sum().reset_index().rename(columns={'상품명': '이름', '수량': '갯수'})
                        inventory.to_excel(writer, sheet_name='재고_값', index=False)
                    output.seek(0)

                    st.success("예측 완료 ✅")
                    st.download_button("▶ 예측 결과 다운로드", output, file_name="예측결과.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.subheader("📊 결과 값 요약")
                    st.dataframe(prediction_summary.style.format(na_rep="-", precision=0), use_container_width=True)

        run_md_prediction()
