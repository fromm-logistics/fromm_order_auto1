import streamlit as st
import pandas as pd
import io
import re
import random
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
from random import randint
from collections import Counter

def to_main():
    st.session_state.page = "main"

def run_album():
    st.button(
        "◀ 메인으로 돌아가기",
        on_click=to_main,
        key="back_to_main"
    )

    # ——————————————————————————————————————————————
    # 스텝 관리 초기화
    # ——————————————————————————————————————————————
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'df_step1' not in st.session_state:
        st.session_state.df_step1 = None
    if 'df_step2' not in st.session_state:
        st.session_state.df_step2 = None

    def go_next():
        st.session_state.step += 1
    def go_prev():
        st.session_state.step = max(1, st.session_state.step - 1)

    # ——————————————————————————————————————————————
    # 상단 UI
    # ——————————————————————————————————————————————
    st.title("📦 앨범 자동화")

    col1, col2, col3 = st.columns([1,1,1])
    with col1:
        st.button("◀ 이전 단계", on_click=go_prev)
    with col2:
        st.markdown(f"### Step {st.session_state.step} / 3")
    with col3:
        st.button("다음 단계 ▶", on_click=go_next)

    st.write("---")

    project_label = st.text_input("▶ 프로젝트 이름", max_chars=50)
    if not project_label:
        st.warning("프로젝트 이름을 입력해주세요.")

    # ——————————————————————————————————————————————
    # STEP 1: 당첨자 제외
    # ——————————————————————————————————————————————
    if st.session_state.step == 1:
        st.header("1️⃣ 당첨자 제외")

        # A 파일 업로드 (원본 주문서)
        file_a = st.file_uploader("A: 원본 CSV 파일 업로드", type="csv", key="step1_a")
        if file_a:
            df_a = pd.read_csv(file_a, dtype={'우편번호': str, '전화번호': str})
            st.success(f"A 파일: {df_a.shape[0]}행 로드 완료")
            st.session_state.df_step1 = df_a.copy()
            st.session_state.df_step1_original = df_a.copy()

        # 재고명 선택 및 B 파일 업로드 (당첨자 명단)
        exclude_stock = None
        file_b = None
        selected_options = []
        
        if st.session_state.df_step1 is not None:
            stocks = sorted(st.session_state.df_step1['재고명'].dropna().unique())
            exclude_stock = st.selectbox("▶ 당첨자 제외할 재고명 선택", stocks)
            file_b = st.file_uploader("B: 당첨자 Excel 파일 업로드", type="xlsx", key="step1_b")

            # B 파일이 업로드되면 옵션 드롭박스 표시
            if file_b:
                df_notice = pd.read_excel(file_b, sheet_name="notice", dtype={'우편번호': str})
                opts = sorted(df_notice["option"].dropna().unique())
                selected_options = st.multiselect("▶ 제외할 옵션(option) 선택", opts)

        # 제외 실행 버튼
        if st.button("🔪 당첨자 제외 실행"):
            if not project_label:
                st.error("프로젝트 이름을 입력해주세요.")
            elif st.session_state.df_step1 is None:
                st.error("먼저 A 파일을 업로드하세요.")
            elif exclude_stock is None:
                st.error("제외할 재고명을 선택하세요.")
            elif file_b is None:
                st.error("B 파일을 업로드하세요.")
            elif not selected_options:
                st.error("제외할 옵션을 하나 이상 선택해주세요.")
            else:
                df = st.session_state.df_step1.copy()
                df_b_full = pd.read_excel(file_b, sheet_name="notice", dtype={'우편번호': str})
                
                # 선택된 옵션에 해당하는 당첨자만 필터링
                df_b_filtered = df_b_full[df_b_full["option"].isin(selected_options)]
                
                # '배송진행여부'가 cancelled인 행 제거
                df = df[df["배송진행여부"] != "cancelled"]
                
                # 필터링된 당첨자 명단의 merchant uid에 대해서만 수량 -1 실행
                for uid in df_b_filtered["merchant uid"].dropna():
                    mask = (df["주문번호"] == uid) & (df["재고명"] == exclude_stock)
                    if mask.any():
                        idx = df[mask].index[0]
                        df.at[idx, "수량"] -= 1
                
                # 수량 0인 행 제거 및 정렬 (1순위: 주문번호, 2순위: id)
                df = df[df["수량"] > 0]
                df["id"] = df["id"].astype(str) # 형충돌 방지
                df = df.sort_values(by=["주문번호", "id"]).reset_index(drop=True)
                
                st.session_state.df_step1 = df
                st.success(f"✅ 선택된 옵션({', '.join(selected_options)}) 당첨자 제외 완료!")
                st.dataframe(df.head())

                # XLSX 결과 생성
                buf_step1 = io.BytesIO()
                with pd.ExcelWriter(buf_step1, engine="openpyxl") as writer:
                    df_a_sorted = st.session_state.get("df_step1_original", pd.DataFrame())
                    df_a_sorted["id"] = df_a_sorted["id"].astype(str)
                    df_a_sorted = df_a_sorted.sort_values(by=["주문번호", "id"])
                    
                    df_a_sorted.to_excel(writer, sheet_name="원본", index=False)
                    df.to_excel(writer, sheet_name="당첨자빼기진행_주문서", index=False)

                buf_step1.seek(0)
                st.session_state["step1_buf"] = buf_step1.getvalue()

        # XLSX 다운로드 버튼
        if st.session_state.df_step1 is not None and "step1_buf" in st.session_state:
            st.download_button(
                "▶ Step1 결과 다운로드 (XLSX)",
                data=st.session_state["step1_buf"],
                file_name=f"{project_label}_당첨자제외_{exclude_stock}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # ――――――――――――――――――――――――――
    # STEP 2: 박스 분할 (재고별 박스 수량 설정 추가)
    # ――――――――――――――――――――――――――
    elif st.session_state.step == 2:
        st.header("2️⃣ 박스 분할")

        up = st.file_uploader("▶ 입력 파일 업로드 (CSV 또는 XLSX)", type=["csv", "xlsx"], key="step2_in")
        if up is None:
            st.stop()

        file_name = up.name.lower()
        dtype_spec = {'우편번호': str, '전화번호': str}

        if file_name.endswith(".csv"):
            df = pd.read_csv(up, dtype=dtype_spec)
        elif file_name.endswith(".xlsx"):
            try:
                df = pd.read_excel(up, sheet_name="당첨자빼기진행_주문서", dtype=dtype_spec)
            except Exception:
                df = pd.read_excel(up, dtype=dtype_spec)
        else:
            st.error("지원하지 않는 파일 형식입니다.")
            st.stop()

        stocks = sorted(df['재고명'].dropna().unique())
        split_stocks = st.multiselect("▶ 박스 분할 대상 재고명 선택 (여러 개 가능)", stocks)

        # 각 재고별 설정을 저장할 딕셔너리
        all_configs = {}

        if split_stocks:
            for s_stock in split_stocks:
                st.write(f"---")
                st.subheader(f"📍 '{s_stock}' 설정")
                
                # 1. 해당 재고의 박스 제한 수량 설정
                b_limit = st.number_input(f"'{s_stock}'의 박스당 최대 수량", min_value=1, value=20, key=f"limit_{s_stock}")
                
                # 2. 매핑 설정
                n_map_key = f"n_map_{s_stock}"
                if n_map_key not in st.session_state:
                    st.session_state[n_map_key] = 5

                cols = st.columns(2)
                current_mapping = {}
                for i in range(st.session_state[n_map_key]):
                    m_name = cols[0].text_input(f"분할될 재고명 #{i+1}", key=f"map_name_{s_stock}_{i}")
                    m_code = cols[1].text_input(f"분할될 재고코드 #{i+1}", key=f"map_code_{s_stock}_{i}")
                    if m_name and m_code:
                        current_mapping[m_name] = m_code
                
                if st.button(f"➕ '{s_stock}' 매핑 추가", key=f"btn_{s_stock}"):
                    st.session_state[n_map_key] += 5
                    st.rerun()
                
                all_configs[s_stock] = {
                    "mapping": current_mapping,
                    "box_limit": b_limit
                }

        st.write("---")

        if st.button("✅ 박스 분할 실행"):
            if not all_configs:
                st.error("분할 설정을 완료해주세요.")
                st.stop()

            if '배송진행여부' in df.columns:
                df = df[df['배송진행여부'] != 'cancelled']

            # 1. 분할 대상(선택된 순서대로)과 비대상 분할
            df_target_full = df[df['재고명'].isin(split_stocks)].copy()
            df_others = df[~df['재고명'].isin(split_stocks)].copy()

            final_list = []

            # 주문번호별로 그룹화하여 처리
            for order_id, group in df_target_full.groupby('주문번호'):
                sorted_group = group.sort_values(by="id")
                
                # [중요] 사용자가 선택한 split_stocks 순서대로 유닛들을 한 줄로 세웁니다.
                all_processed_units = []
                for s_stock in split_stocks:
                    stock_group = sorted_group[sorted_group['재고명'] == s_stock]
                    if stock_group.empty: continue
                    
                    config = all_configs[s_stock]
                    mapping = config["mapping"]
                    limit = config["box_limit"]
                    p_names = list(mapping.keys())
                    
                    # 유닛 전개 및 순환 이름 배정
                    temp_units = []
                    for _, row in stock_group.iterrows():
                        for _ in range(int(row['수량'])):
                            temp_units.append(row.copy())
                    
                    for idx, unit in enumerate(temp_units):
                        assigned_name = p_names[idx % len(p_names)]
                        unit['재고명'] = assigned_name
                        unit['재고코드'] = mapping[assigned_name]
                        # 나중에 계산을 위해 이 재고가 차지하는 '비중'을 기록합니다.
                        unit['_weight'] = 1.0 / limit 
                        all_processed_units.append(unit)

                # 2. '잔여 공간' 기반 박스 배분 로직
                order_boxes = []
                current_box = []
                current_weight_sum = 0.0
                
                for unit in all_processed_units:
                    unit_weight = unit['_weight']
                    
                    # 만약 현재 유닛을 넣었을 때 1.0(박스 가득참)을 초과한다면
                    if current_weight_sum + unit_weight > 1.0000000001:
                        order_boxes.append(current_box)
                        current_box = [unit]
                        current_weight_sum = unit_weight
                    else:
                        current_box.append(unit)
                        current_weight_sum += unit_weight
                
                if current_box:
                    order_boxes.append(current_box)

                # 3. 결과 리스트 생성 (박스 번호 부여 및 집계)
                for box_num, box_content in enumerate(order_boxes, start=1):
                    temp_counts = defaultdict(int)
                    id_to_sample_row = {}
                    
                    for unit_row in box_content:
                        id_val = str(unit_row.get('id', ''))
                        p_name = unit_row['재고명']
                        key = (id_val, p_name)
                        temp_counts[key] += 1
                        id_to_sample_row[(id_val, p_name)] = unit_row
                    
                    for (id_val, p_name), qty in temp_counts.items():
                        new_row = id_to_sample_row[(id_val, p_name)].copy()
                        new_row['주문번호'] = f"{order_id}-{box_num}"
                        new_row['재고명'] = p_name
                        new_row['재고코드'] = new_row['재고코드']
                        new_row['수량'] = qty
                        if '_weight' in new_row: del new_row['_weight']
                        final_list.append(new_row)

            # 4. 비대상 주문 처리
            for _, row in df_others.iterrows():
                r = row.copy()
                if '-' not in str(r['주문번호']):
                    r['주문번호'] = f"{r['주문번호']}-1"
                final_list.append(r)

            df_full = pd.DataFrame(final_list)

            # 후처리 로직
            df_full['상품명'] = df_full['재고명']
            df_full['옵션명'] = df_full['재고명']
            df_full['id'] = df_full['id'].astype(str)
            df_full = df_full.sort_values(by=["주문번호", "id"]).reset_index(drop=True)

            # 공통 후처리 로직
            df_full['상품금액'] = df_full['상품금액'].replace(0, 1)
            mask = (df_full['국가코드'] != 'KR') & (df_full['결제통화'] == 'KRW')
            df_full.loc[mask, '결제통화'] = 'USD'
            df_full.loc[mask & (df_full['상품금액'] > 1), '상품금액'] /= 1000
            df_full['실결제금액'] = df_full['수량'] * df_full['상품금액']
            
            if '실결제금액' in df_full.columns:
                col_idx = df_full.columns.get_loc('상품금액') + 1
                cols = df_full.columns.tolist()
                cols.insert(col_idx, cols.pop(cols.index('실결제금액')))
                df_full = df_full[cols]

            df_dom = df_full[df_full["국가코드"] == "KR"].copy()
            df_int = df_full[df_full["국가코드"] != "KR"].copy()

            def format_phone_number(phone):
                phone = re.sub(r'\D', '', str(phone))
                if len(phone) == 11:
                    return f"{phone[:3]}-{phone[3:7]}-{phone[7:]}"
                elif len(phone) == 10:
                    return f"{phone[:3]}-{phone[3:6]}-{phone[6:]}"
                return phone
            df_dom['전화번호'] = df_dom['전화번호'].map(format_phone_number)

            df_int['희망배송사'] = df_int['국가코드'].map(
                lambda c: 'sagawa' if c == 'JP' else ('emspremium' if c in ['US', 'IT', 'CO', 'RO', 'DE', 'NL', 'SE'] else 'ems')
            )
            jp_mask = df_int["국가코드"] == "JP"
            df_int.loc[jp_mask, "도로명주소"] = (
                df_int.loc[jp_mask, "도로명주소"].fillna('') + ' ' +
                df_int.loc[jp_mask, "상세주소"].fillna('') + ' ' +
                df_int.loc[jp_mask, "도시"].fillna('') + ' ' +
                df_int.loc[jp_mask, "주"].fillna('')
            )
            df_int.loc[jp_mask, ['상세주소', '주']] = ''
            df_int.loc[jp_mask, ['도시']] = '.'
            df_int.loc[df_int["국가명"] == "Japan", "국가명"] = "."

            def generate_item_code(row, index):
                if pd.isna(row['id']) or row['id'] == '' or row['id'] == 'nan':
                    rnd = f"{random.randint(0,9999999999):010d}"
                    return f"NAN_{rnd[:3]}_{rnd[3:6]}_{rnd[6:]}"
                shipping_code = str(row['배송비묶음'])[:3] if pd.notnull(row['배송비묶음']) else "000"
                return f"{row['id']}-{shipping_code}-{index + 1}"

            df_int['판매사 품주번호'] = [generate_item_code(r, i) for i, r in df_int.iterrows()]

            inv = df_full.groupby('상품명', as_index=False)['수량'].sum().rename(columns={'상품명': '이름', '수량': '갯수'})

            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                df.sort_values(by=["주문번호", "id"], ascending=True).to_excel(writer, sheet_name="원본", index=False)
                df_dom.to_excel(writer, sheet_name="국내", index=False)
                df_int.to_excel(writer, sheet_name="해외", index=False)
                inv.to_excel(writer, sheet_name="재고_값", index=False)

            wb = load_workbook(buf)
            yellow = PatternFill("solid", fgColor="FFFF00")
            
            # 하이라이트 적용
            ws_dom = wb["국내"]
            zc = df_dom.columns.get_loc("우편번호") + 1
            ph = df_dom.columns.get_loc("전화번호") + 1
            for i, row in enumerate(df_dom.itertuples(), start=2):
                if len(str(row.우편번호)) != 5:
                    ws_dom.cell(i, zc).fill = yellow
                if not str(row.전화번호).startswith("010"):
                    ws_dom.cell(i, ph).fill = yellow

            ws_int = wb["해외"]
            name_idx = df_int.columns.get_loc("받는사람") + 1
            zip_idx = df_int.columns.get_loc("우편번호") + 1
            cc_idx = df_int.columns.get_loc("국가코드") + 1
            for i, row in enumerate(df_int.itertuples(), start=2):
                if str(row.받는사람).isdigit():
                    ws_int.cell(i, name_idx).fill = yellow
                if row.국가코드 == 'HK':
                    ws_int.cell(i, cc_idx).fill = yellow
                    if row.우편번호 in {'-', '0', '00', '000', '0000', '00000', '000000'}:
                        ws_int.cell(i, zip_idx, value='999077')

            wb.save(buf)
            st.success("✅ 박스 분할 및 데이터 업데이트 완료!")
            st.download_button("📥 엑셀 다운로드", data=buf.getvalue(), file_name=f"{project_label}.박스분할결과.xlsx")

            col_d, col_i = st.columns(2)
            col_d.metric("국내 주문번호 수 (중복제외)", df_dom['주문번호'].nunique())
            col_i.metric("해외 주문번호 수 (중복제외)", df_int['주문번호'].nunique())
            

    # ——————————————————————————————————————————————
    # STEP 3: 당첨자 추가
    # ——————————————————————————————————————————————
    else:
        st.header("3️⃣ 당첨자 추가")
        df_domestic = df_international = None
        up2 = st.file_uploader("▶ 입력 XLSX (Step2 결과)", type="xlsx", key="step3_in")
        if up2:
            dtype_spec = {'우편번호': str, '전화번호': str}
            df_domestic = pd.read_excel(up2, sheet_name="국내", dtype=dtype_spec)
            df_international = pd.read_excel(up2, sheet_name="해외", dtype=dtype_spec)
            df_orig = pd.read_excel(up2, sheet_name="원본", dtype=dtype_spec)
        else:
            st.warning("Step2 결과를 업로드해주세요.")
            st.stop()

        file_b2 = st.file_uploader("▶ 당첨자 추가용 Excel 파일(B)", type="xlsx", key="step3_b")
        df_notice = None
        selected_options = []
        if file_b2:
            df_notice = pd.read_excel(file_b2, sheet_name="notice", dtype={'우편번호': str})
            opts = sorted(df_notice["option"].dropna().unique())
            selected_options = st.multiselect("▶ 추가할 옵션(option) 선택", opts)

        custom = {"상품명": st.text_input("▶ 추가할 상품명"), "재고코드": st.text_input("▶ 추가 재고코드"), "수량": 1, "상품금액": 1, "결제통화": "USD"}
        custom["옵션명"] = custom["상품명"]

        if st.button("✅ 당첨자 추가 실행"):
            try:
                df_sel = df_notice[df_notice["option"].isin(selected_options)]
                uid_counts = df_sel["merchant uid"].dropna().value_counts().to_dict()
                dom = df_domestic.copy()
                intl = df_international.copy()
                dom["원주문번호"] = dom["주문번호"].str[:20]
                intl["원주문번호"] = intl["주문번호"].str[:20]

                add_dom = dom[dom["원주문번호"].isin(uid_counts)].drop_duplicates("원주문번호").copy()
                add_intl = intl[intl["원주문번호"].isin(uid_counts)].drop_duplicates("원주문번호").copy()

                for df_add in (add_dom, add_intl):
                    df_add["id"] = ""
                    df_add["수량"] = df_add["원주문번호"].map(lambda wn: custom["수량"] * uid_counts.get(wn, 0))
                    df_add["상품명"] = custom["상품명"]
                    df_add["옵션명"] = custom["옵션명"]
                    df_add["재고코드"] = custom["재고코드"]
                    df_add["상품금액"] = custom["상품금액"]
                    df_add["결제통화"] = custom["결제통화"]
                    df_add["재고명"] = custom["상품명"]
                    df_add["실결제금액"] = df_add["수량"] * df_add["상품금액"]

                # 합본 생성
                dom_all = pd.concat([dom.drop(columns=["원주문번호"]), add_dom.drop(columns=["원주문번호"], errors="ignore")], ignore_index=True)
                intl_all = pd.concat([intl.drop(columns=["원주문번호"]), add_intl.drop(columns=["원주문번호"], errors="ignore")], ignore_index=True)

                # 정렬 전 모든 id 컬럼을 문자열로 강제 변환
                dom_all["id"] = dom_all["id"].astype(str)
                intl_all["id"] = intl_all["id"].astype(str)
                add_dom["id"] = add_dom["id"].astype(str)
                add_intl["id"] = add_intl["id"].astype(str)
                df_orig["id"] = df_orig["id"].astype(str)

                # 최종 합본 정렬
                dom_all = dom_all.sort_values(by=["주문번호", "id"]).reset_index(drop=True)
                intl_all = intl_all.sort_values(by=["주문번호", "id"]).reset_index(drop=True)
                add_dom = add_dom.sort_values(by=["주문번호", "id"]).reset_index(drop=True)
                add_intl = add_intl.sort_values(by=["주문번호", "id"]).reset_index(drop=True)

                st.session_state["step3_dom_count"] = dom_all['주문번호'].nunique()
                st.session_state["step3_int_count"] = intl_all['주문번호'].nunique()

                buf_all = io.BytesIO()
                with pd.ExcelWriter(buf_all, engine="openpyxl") as w:
                    df_orig.sort_values(by=["주문번호", "id"]).to_excel(w, sheet_name="원본", index=False)
                    df_notice.to_excel(w, sheet_name="notice", index=False)
                    add_dom.to_excel(w, sheet_name="국내_추가", index=False)
                    add_intl.to_excel(w, sheet_name="해외_추가", index=False)
                    dom_all.to_excel(w, sheet_name="국내", index=False)
                    intl_all.to_excel(w, sheet_name="해외", index=False)
                
                buf_all.seek(0)
                st.session_state["step3_buf_all"] = buf_all.getvalue()
                st.success("✅ 당첨자 추가 및 주문번호/id 기준 정렬 완료!")
            except Exception:
                st.error(traceback.format_exc())

    if "step3_buf_all" in st.session_state:
        st.download_button("▶ 전체 결과 다운로드", data=st.session_state["step3_buf_all"], file_name=f"{project_label}_최종본_정렬.xlsx")
        if "step3_dom_count" in st.session_state:
            col_d, col_i = st.columns(2)
            col_d.metric("국내 주문번호 수 (중복제외)", st.session_state["step3_dom_count"])
            col_i.metric("해외 주문번호 수 (중복제외)", st.session_state["step3_int_count"])
